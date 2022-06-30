
from openpyxl import load_workbook



class Students:
    def __init__(self, name=None, adno=None, phy=None, che=None, maths=None):
        self.name = name
        self.adno = adno
        self.phy = phy
        self.che = che
        self.maths = maths

    def  percFn(self):
        perc = self.totalMarks() / 3
        return perc

    def  totalMarks(self):
        total = self.phy + self.che + self.maths
        return total

    def  phyGrd(self):
        if (self.phy > 90 and self.phy <= 100):
            print("      Grade:  A1")
            print("      Grade Point: 10.0")
        elif (self.phy > 80 and self.phy <= 90):
            print("      Grade:  A2")
            print("      Grade Point: 9.0")
        elif (self.phy > 70 and self.phy <= 80):
            print("      Grade:  B1")
            print("      Grade Point: 8.0")
        elif (self.phy > 60 and self.phy <= 70):
            print("      Grade:   B2")
            print("      Grade Point: 7.0")
        elif (self.phy > 50 and self.phy <= 60):
            print("      Grade:  C1")
            print("      Grade Point: 6.0")
        elif (self.phy > 40 and self.phy <= 50):
            print("      Grade:  C2")
            print("      Grade Point: 5.0")
        elif (self.phy > 32 and self.phy <= 40):
            print("      Grade:  D")
            print("      Grade Point: 4.0")
        elif (self.phy > 20 and self.phy <= 32):
            print("      Grade:  E1")
            print("      Grade Point: C")
        elif (self.phy >= 00 and self.phy <= 20):
            print("      Grade:  E2")
            print("      Grade Point: C")
        else:
            print("No data")



    def  cheGrd(self):
        if (self.che > 90 and self.che <= 100):
            print("      Grade:  A1")
            print("      Grade Point: 10.0")
        elif (self.che > 80 and self.che <= 90):
            print("      Grade:  A2")
            print("      Grade Point: 9.0")
        elif (self.che > 70 and self.che <= 80):
            print("      Grade:  B1")
            print("      Grade Point: 8.0")
        elif (self.che > 60 and self.che <= 70):
            print("      Grade:  B2")
            print("      Grade Point: 7.0")
        elif (self.che > 50 and self.che <= 60):
            print("      Grade:  C1")
            print("      Grade Point: 6.0")
        elif (self.che > 40 and self.che <= 50):
            print("      Grade:  C2")
            print("      Grade Point: 5.0")
        elif (self.che > 32 and self.che <= 40):
            print("      Grade:  D")
            print("      Grade Point: 4.0")
        elif (self.che > 20 and self.che <= 32):
            print("      Grade:  E1")
            print("      Grade Point: C")
        elif (self.che >= 00 and self.che <= 20):
            print("      Grade:  E2")
            print("      Grade Point: C")
        else:
            print("No data")



    def  mathsGrd(self):
        if (self.maths > 90 and self.maths <= 100):
            print("      Grade:  A1")
            print("      Grade Point: 10.0")
        elif (self.maths > 80 and self.maths <= 90):
            print("      Grade:  A2")
            print("      Grade Point: 9.0")
        elif (self.maths > 70 and self.maths <= 80):
            print("      Grade:  B1")
            print("      Grade Point: 8.0")
        elif (self.maths > 60 and self.maths <= 70):
            print("      Grade:  B2")
            print("      Grade Point: 7.0")
        elif (self.maths > 50 and self.maths <= 60):
            print("      Grade:  C1")
            print("      Grade Point: 6.0")
        elif (self.maths > 40 and self.maths <= 50):
            print("      Grade:  C2")
            print("      Grade Point: 5.0")
        elif (self.maths > 32 and self.maths <= 40):
            print("      Grade:  D")
            print("      Grade Point: 4.0")
        elif (self.maths > 20 and self.maths <= 32):
            print("      Grade:  E1")
            print("      Grade Point: C")
        elif (self.maths >= 00 and self.maths <= 20):
            print("      Grade:  E2")
            print("      Grade Point: C")
        else:
            print("No data")
            

wb = load_workbook('E:\your-file-location\data.xlsx')
sheetname = "Data"
ws = wb[sheetname]


header = [cell.value for cell in ws[1]]

st = []

for row in list(ws.rows)[1:]:
    args = [cell.value for cell in row]
    person = Students(*args)
    st.append(person)
        
                                            
   

         


print("WELCOME TO STUDENT DATA DIRECTORY")

while True:
    print("\nMAIN MENU")
    print("1. Name of the Student")
    print("2. Admission number of the Student")
    print("3. Exit")
    choice = int(input("Enter the Choice:"))

    if choice == 1:
        
        nam = input("Enter Name:")
        
        for i in range(0,ws.max_row-1):
            if nam == st[i].name:
                print("Name:"+ st[i].name)
                print("Admission No:", st[i].adno)
                per = st[i].percFn()
                print("Percentage:", per)
                print("Physics: \n "
                          "     Mark:", st[i].phy )
                st[i].phyGrd()
                print("Chemistry: \n "
                  "     Mark:", st[i].che)
                st[i].cheGrd()
                print("Maths: \n "
                  "     Mark:", st[i].maths)
                st[i].mathsGrd()

           
        
   

    elif choice == 2:
        ad = int(input("Enter Admission No:"))
        
        for i in range(0,ws.max_row-1):
            if ad == st[i].adno:
                print("Name:"+ st[i].name)
                print("Admission No:", st[i].adno)
                per = st[i].percFn()
                print("Percentage:", per)
                print("Physics: \n "
                          "     Mark:", st[i].phy )
                st[i].phyGrd()
                print("Chemistry: \n "
                  "     Mark:", st[i].che)
                st[i].cheGrd()
                print("Maths: \n "
                  "     Mark:", st[i].maths)
                st[i].mathsGrd()


        


    elif choice == 3:
        break

    else:
        print("Oops!Incorrect choice")





